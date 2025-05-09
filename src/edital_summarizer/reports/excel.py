import pandas as pd
from typing import List, Dict, Any
import os
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class ExcelReportGenerator:
    """
    Gera relatórios em formato Excel a partir dos dados processados dos editais.
    """

    def generate_report(self, results: List[Dict[str, Any]], output_file: str) -> str:
        """
        Gera um relatório Excel com os resultados do processamento.

        Args:
            results: Lista de dicionários com os resultados do processamento
            output_file: Caminho para o arquivo Excel de saída

        Returns:
            Caminho do arquivo Excel gerado
        """
        # Criar um novo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumo de Editais"

        # Definir cabeçalhos
        headers = [
            "Nome do Arquivo",
            "Tipo do Documento",
            "Propósito",
            "Número do Edital",
            "Número do Processo",
            "Número da Licitação",
            "Órgão",
            "Objeto",
            "Telefone",
            "Website",
            "Local",
            "Resumo Executivo",
            "Resumo Técnico",
        ]

        # Adicionar cabeçalhos
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Adicionar dados
        for result in results:
            print("\n=== DEBUG: Processando resultado para Excel ===")
            print(f"Tipo do resultado: {type(result)}")
            print(f"Chaves disponíveis: {result.keys()}")
            print(f"Conteúdo do resultado: {result}")

            # Extrair metadados
            metadata = result.get("metadata", {})
            identifier = metadata.get("identifier", {})
            organization = metadata.get("organization", {})
            subject = metadata.get("subject", {})

            print("\n=== DEBUG: Metadados extraídos ===")
            print(f"Metadata: {metadata}")
            print(f"Identifier: {identifier}")
            print(f"Organization: {organization}")
            print(f"Subject: {subject}")

            # Preparar dados da linha
            row_data = [
                result.get("file_name", ""),
                result.get("document_type", "Não identificado"),
                result.get("document_purpose", "Não identificado"),
                metadata.get("identifier", {}).get("public_notice", ""),
                metadata.get("identifier", {}).get("process_id", ""),
                metadata.get("identifier", {}).get("bid_number", ""),
                metadata.get("organization", {}).get("organization", ""),
                metadata.get("subject", {}).get("object", ""),
                metadata.get("organization", {}).get("phone", ""),
                metadata.get("organization", {}).get("website", ""),
                metadata.get("organization", {}).get("location", ""),
            ]

            print("\n=== DEBUG: Dados da linha preparados ===")
            print(f"Row data: {row_data}")

            # Formatar os resumos
            executive_summary = result.get('executive_summary', '')
            technical_summary = result.get('technical_summary', '')

            # Remover texto em inglês do início do resumo executivo
            if executive_summary.lower().startswith('given the'):
                executive_summary = executive_summary[executive_summary.find('\n'):].strip()

            # Limpar formatação markdown e caracteres especiais
            executive_summary = executive_summary.replace('#', '').replace('*', '').strip()
            technical_summary = technical_summary.replace('#', '').replace('*', '').strip()

            # Remover cabeçalhos redundantes
            executive_summary = executive_summary.replace('RESUMO EXECUTIVO:', '').replace('RESUMO EXECUTIVO', '').strip()
            technical_summary = technical_summary.replace('RESUMO TÉCNICO:', '').replace('RESUMO TÉCNICO', '').strip()

            # Garantir que os resumos estejam separados
            if executive_summary and technical_summary:
                executive_summary = f"RESUMO EXECUTIVO:\n{executive_summary}"
                technical_summary = f"\n\nRESUMO TÉCNICO:\n{technical_summary}"

            # Preparar os dados da linha
            row_data.extend([executive_summary, technical_summary])

            # Adicionar linha ao worksheet
            ws.append(row_data)

            # Ajustar largura das colunas
            for idx, col in enumerate(ws.columns, 1):
                max_length = 0
                column = get_column_letter(idx)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = min(adjusted_width, 100)  # Limitar largura máxima

        # Salvar o arquivo
        wb.save(output_file)

        # Imprimir o conteúdo do relatório
        print("\nConteúdo do Relatório Excel:")
        print("-" * 100)
        print("\t".join(headers))
        for row in ws.iter_rows(min_row=2, values_only=True):
            print("\t".join(str(cell) if cell is not None else "" for cell in row))
        print("-" * 100)

        return output_file
